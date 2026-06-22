# ─── generate_report.py — .xlsx Test Report Generator ─────────────────────────
"""
Parses JUnit XML output from pytest and generates a styled .xlsx report
matching the SocialShield E2E test report format.

Usage:
    python tests/generate_report.py

Reads:  test-results/results.xml
Writes: test-results/E2E_Test_Report_SocialShield_<timestamp>.xlsx
"""
import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Configuration ────────────────────────────────────────────────────────────
RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "test-results")
XML_PATH = os.path.join(RESULTS_DIR, "results.xml")

# Test case metadata mapping — maps test function names to details
TEST_METADATA = {
    # ── Splash Page ──
    "test_tc001_page_loads_with_correct_title":     {"id": "TC-001", "category": "Functionality", "page": "Splash Page",     "severity": "High",   "description": "Verify the page loads and has the correct document title", "input": "Navigate to base URL", "expected": "Page title contains 'SocialShield'"},
    "test_tc002_shield_logo_visible":               {"id": "TC-002", "category": "Functionality", "page": "Splash Page",     "severity": "Medium", "description": "Verify the shield logo emoji (🛡️) is visible on splash", "input": "Navigate to splash page", "expected": "Shield logo element is visible"},
    "test_tc003_brand_text_displayed":               {"id": "TC-003", "category": "Functionality", "page": "Splash Page",     "severity": "High",   "description": "Verify 'SocialShield' brand text is displayed", "input": "Navigate to splash page", "expected": "Brand text 'SocialShield' visible"},
    "test_tc004_subtitle_contains_ai_powered":       {"id": "TC-004", "category": "Functionality", "page": "Splash Page",     "severity": "Medium", "description": "Verify subtitle contains 'AI-Powered' text", "input": "Navigate to splash page", "expected": "Subtitle with 'AI-Powered' is present"},
    "test_tc005_auto_redirect_within_timeout":       {"id": "TC-005", "category": "Functionality", "page": "Splash Page",     "severity": "Critical","description": "Verify splash auto-redirects within 4 seconds", "input": "Navigate to splash, wait 4s", "expected": "URL changes from splash to another page"},
    "test_tc006_dot_indicators_present":             {"id": "TC-006", "category": "Functionality", "page": "Splash Page",     "severity": "Low",    "description": "Verify 3 dot/bar indicators are rendered on splash", "input": "Navigate to splash page", "expected": "At least 3 dot indicators are present"},

    # ── Onboarding Page ──
    "test_tc007_first_slide_shows_verify_reality":   {"id": "TC-007", "category": "Functionality", "page": "Onboarding Page", "severity": "High",   "description": "Verify first slide shows 'Verify Reality' title", "input": "Navigate to /onboarding", "expected": "'Verify Reality' text is visible"},
    "test_tc008_next_button_advances_slide":         {"id": "TC-008", "category": "Functionality", "page": "Onboarding Page", "severity": "High",   "description": "Verify Next button advances to the second slide", "input": "Click Next button", "expected": "Slide 2 'AI-Powered Analysis' is shown"},
    "test_tc009_four_dot_indicators_render":          {"id": "TC-009", "category": "Functionality", "page": "Onboarding Page", "severity": "Medium", "description": "Verify all 4 dot indicators render for 4 slides", "input": "Navigate to /onboarding", "expected": "4 dot navigation buttons present"},
    "test_tc010_clicking_dot_navigates_to_slide":    {"id": "TC-010", "category": "Functionality", "page": "Onboarding Page", "severity": "Medium", "description": "Verify clicking a dot navigates to the correct slide", "input": "Click 3rd dot indicator", "expected": "Slide 3 'Scan Anything' is shown"},
    "test_tc011_skip_button_navigates_to_auth":      {"id": "TC-011", "category": "Functionality", "page": "Onboarding Page", "severity": "High",   "description": "Verify Skip button navigates to auth page", "input": "Click Skip button", "expected": "Redirected to /auth"},
    "test_tc012_last_slide_shows_get_started":        {"id": "TC-012", "category": "Functionality", "page": "Onboarding Page", "severity": "High",   "description": "Verify last slide shows 'Get Started' button", "input": "Navigate to slide 4", "expected": "'Get Started' button visible"},
    "test_tc013_get_started_navigates_to_auth":      {"id": "TC-013", "category": "Functionality", "page": "Onboarding Page", "severity": "High",   "description": "Verify 'Get Started' button navigates to auth page", "input": "Click 'Get Started' on last slide", "expected": "Redirected to /auth"},
    "test_tc014_slide_icons_change_per_slide":       {"id": "TC-014", "category": "Functionality", "page": "Onboarding Page", "severity": "Low",    "description": "Verify slide icons change when navigating between slides", "input": "Navigate between slides", "expected": "Page content changes per slide"},
    "test_tc015_subtitle_updates_per_slide":          {"id": "TC-015", "category": "Functionality", "page": "Onboarding Page", "severity": "Low",    "description": "Verify subtitle text updates for each slide", "input": "Navigate to slide 1", "expected": "Subtitle mentions deepfake detection"},
    "test_tc016_onboarded_flag_set_after_finish":     {"id": "TC-016", "category": "Functionality", "page": "Onboarding Page", "severity": "Critical","description": "Verify 'ss_onboarded' localStorage flag is set after onboarding", "input": "Complete onboarding flow", "expected": "ss_onboarded = '1' in localStorage"},

    # ── Auth Page ──
    "test_tc017_auth_card_renders_with_logo":         {"id": "TC-017", "category": "Functionality", "page": "Auth Page",       "severity": "High",   "description": "Verify auth card renders with the SocialShield logo", "input": "Navigate to /auth", "expected": "Auth card and logo elements present"},
    "test_tc018_welcome_back_heading_login_mode":    {"id": "TC-018", "category": "Functionality", "page": "Auth Page",       "severity": "High",   "description": "Verify 'Welcome Back' heading in login mode", "input": "Navigate to /auth", "expected": "'Welcome Back' heading visible"},
    "test_tc019_create_account_heading_signup_mode":  {"id": "TC-019", "category": "Functionality", "page": "Auth Page",       "severity": "High",   "description": "Verify 'Create Account' heading when toggled to signup", "input": "Click 'Sign Up' toggle", "expected": "'Create Account' heading visible"},
    "test_tc020_toggle_between_signin_and_signup":   {"id": "TC-020", "category": "Functionality", "page": "Auth Page",       "severity": "High",   "description": "Verify toggle between Sign In ↔ Sign Up modes", "input": "Toggle auth mode twice", "expected": "Heading switches correctly both ways"},
    "test_tc021_email_input_present_and_functional": {"id": "TC-021", "category": "Functionality", "page": "Auth Page",       "severity": "Critical","description": "Verify email input field accepts input", "input": "Type test@example.com", "expected": "Input value equals typed text"},
    "test_tc022_password_input_present":              {"id": "TC-022", "category": "Functionality", "page": "Auth Page",       "severity": "Critical","description": "Verify password input field is present with type 'password'", "input": "Navigate to /auth", "expected": "Password input with type='password' exists"},
    "test_tc023_password_visibility_toggle":          {"id": "TC-023", "category": "Functionality", "page": "Auth Page",       "severity": "Medium", "description": "Verify password visibility toggle (👁️ ↔ 🙈)", "input": "Click toggle twice", "expected": "Type toggles: password → text → password"},
    "test_tc024_submit_button_present":               {"id": "TC-024", "category": "Functionality", "page": "Auth Page",       "severity": "Critical","description": "Verify submit button is present with type='submit'", "input": "Navigate to /auth", "expected": "Submit button exists"},
    "test_tc025_empty_form_submission_shows_error":   {"id": "TC-025", "category": "Functionality", "page": "Auth Page",       "severity": "High",   "description": "Verify submitting empty form shows error message", "input": "Click submit with empty fields", "expected": "Error message about required fields"},
    "test_tc026_short_password_validation":            {"id": "TC-026", "category": "Functionality", "page": "Auth Page",       "severity": "High",   "description": "Verify short password (<6 chars) shows validation error", "input": "Enter email + 3-char password, submit", "expected": "Error about 6 character minimum"},
    "test_tc027_google_signin_button_present":        {"id": "TC-027", "category": "Functionality", "page": "Auth Page",       "severity": "High",   "description": "Verify Google Sign-In button is present", "input": "Navigate to /auth", "expected": "Google button with 'Google' text exists"},
    "test_tc028_divider_between_form_and_google":     {"id": "TC-028", "category": "Functionality", "page": "Auth Page",       "severity": "Low",    "description": "Verify 'or' divider between form and Google button", "input": "Navigate to /auth", "expected": "Divider element with 'or' text"},
    "test_tc029_error_clears_when_toggling_mode":     {"id": "TC-029", "category": "Functionality", "page": "Auth Page",       "severity": "Medium", "description": "Verify error message clears when toggling auth mode", "input": "Trigger error, then toggle mode", "expected": "Error message disappears"},
    "test_tc030_form_inputs_have_autocomplete":       {"id": "TC-030", "category": "Functionality", "page": "Auth Page",       "severity": "Medium", "description": "Verify form inputs have correct autocomplete attributes", "input": "Inspect email and password inputs", "expected": "autocomplete='email' and 'current-password'"},

    # ── Home Page ──
    "test_tc031_page_header_shows_welcome_back":     {"id": "TC-031", "category": "Functionality", "page": "Home Page",       "severity": "High",   "description": "Verify page header shows 'Welcome back' text", "input": "Navigate to /home (authenticated)", "expected": "'Welcome back' text visible"},
    "test_tc032_username_greeting_displayed":          {"id": "TC-032", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify username/greeting is displayed", "input": "Navigate to /home", "expected": "User display name shown"},
    "test_tc033_trust_score_card_renders":             {"id": "TC-033", "category": "Functionality", "page": "Home Page",       "severity": "High",   "description": "Verify AI Trust Score card renders", "input": "Navigate to /home", "expected": "Trust Score card element present"},
    "test_tc034_trust_score_shows_value":              {"id": "TC-034", "category": "Functionality", "page": "Home Page",       "severity": "High",   "description": "Verify trust score shows a value out of /100", "input": "Navigate to /home", "expected": "'/100' format displayed"},
    "test_tc035_three_stat_cards_render":              {"id": "TC-035", "category": "Functionality", "page": "Home Page",       "severity": "High",   "description": "Verify 3 stat cards render (Total Scans, Fake Detected, Suspicious)", "input": "Navigate to /home", "expected": "≥3 stat-card elements present"},
    "test_tc036_scan_detect_heading_present":          {"id": "TC-036", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify 'Scan & Detect' section heading is present", "input": "Navigate to /home", "expected": "'Scan & Detect' text visible"},
    "test_tc037_six_scan_type_cards_render":           {"id": "TC-037", "category": "Functionality", "page": "Home Page",       "severity": "Critical","description": "Verify all 6 scan type cards render", "input": "Navigate to /home", "expected": "6 scan-type-card elements present"},
    "test_tc038_image_scan_card_correct":              {"id": "TC-038", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify Image scan card has correct label", "input": "Find #scan-image", "expected": "'Scan Image' text present"},
    "test_tc039_video_scan_card_navigates":            {"id": "TC-039", "category": "Functionality", "page": "Home Page",       "severity": "High",   "description": "Verify Video scan card click navigates to /scan/video", "input": "Click #scan-video", "expected": "URL contains /scan/video"},
    "test_tc040_audio_scan_card_description":          {"id": "TC-040", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify Audio scan card displays sub-description", "input": "Find #scan-audio", "expected": "'Voice clone' text present"},
    "test_tc041_text_scan_card_present":               {"id": "TC-041", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify Text scan card is present", "input": "Find #scan-text", "expected": "'Scan Text' text present"},
    "test_tc042_url_scan_card_clickable":              {"id": "TC-042", "category": "Functionality", "page": "Home Page",       "severity": "High",   "description": "Verify URL scan card is clickable and navigates", "input": "Click #scan-url", "expected": "URL contains /scan/url"},
    "test_tc043_profile_scan_card_renders":            {"id": "TC-043", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify Profile scan card renders", "input": "Find #scan-profile", "expected": "'Scan Profile' text present"},
    "test_tc044_recent_scans_section_displays":        {"id": "TC-044", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify Recent Scans section renders with demo data", "input": "Navigate to /home", "expected": "'Recent Scans' heading or scan data"},
    "test_tc045_view_all_link_navigates_to_history":  {"id": "TC-045", "category": "Functionality", "page": "Home Page",       "severity": "Medium", "description": "Verify 'View All →' link navigates to history", "input": "Click 'View All' link", "expected": "URL contains /history"},
    "test_tc046_scan_cards_have_hover_indicator":     {"id": "TC-046", "category": "Functionality", "page": "Home Page",       "severity": "Low",    "description": "Verify scan type cards have 'Tap to scan' indicator", "input": "Inspect scan card text", "expected": "'Tap to scan' text present"},

    # ── Scan Page ──
    "test_tc047_image_scan_page_loads":                {"id": "TC-047", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify Image scan page loads with correct header", "input": "Navigate to /scan/image", "expected": "'Image' and 'Detection' in header"},
    "test_tc048_back_button_functional":               {"id": "TC-048", "category": "Functionality", "page": "Scan Page",       "severity": "Medium", "description": "Verify back button (←) is present and functional", "input": "Navigate to /scan/image", "expected": "Back button element exists"},
    "test_tc049_upload_zone_renders_for_image":       {"id": "TC-049", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify file upload zone renders for image type", "input": "Navigate to /scan/image", "expected": "upload-zone element present"},
    "test_tc050_upload_zone_file_type_hint":           {"id": "TC-050", "category": "Functionality", "page": "Scan Page",       "severity": "Medium", "description": "Verify upload zone shows file type hints (JPG, PNG, WEBP)", "input": "Navigate to /scan/image", "expected": "File type text visible"},
    "test_tc051_video_scan_page_header":               {"id": "TC-051", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify Video scan page shows 'Video Deepfake Detection'", "input": "Navigate to /scan/video", "expected": "'Video' and 'Detection' in text"},
    "test_tc052_audio_scan_page_loads":                {"id": "TC-052", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify Audio scan page loads correctly", "input": "Navigate to /scan/audio", "expected": "'Voice' or 'Audio' in text"},
    "test_tc053_text_scan_has_textarea":               {"id": "TC-053", "category": "Functionality", "page": "Scan Page",       "severity": "Critical","description": "Verify Text scan page has textarea input", "input": "Navigate to /scan/text", "expected": "#text-input textarea exists"},
    "test_tc054_text_input_placeholder":               {"id": "TC-054", "category": "Functionality", "page": "Scan Page",       "severity": "Medium", "description": "Verify Text input placeholder contains 'suspicious message'", "input": "Inspect #text-input placeholder", "expected": "Placeholder has 'suspicious'"},
    "test_tc055_text_character_counter":               {"id": "TC-055", "category": "Functionality", "page": "Scan Page",       "severity": "Low",    "description": "Verify text character counter updates on typing", "input": "Type 'Hello test' in textarea", "expected": "Counter shows '10' or '/ 10,000'"},
    "test_tc056_url_scan_has_input":                   {"id": "TC-056", "category": "Functionality", "page": "Scan Page",       "severity": "Critical","description": "Verify URL scan page has URL input field", "input": "Navigate to /scan/url", "expected": "#url-input element exists"},
    "test_tc057_url_input_type_is_url":                {"id": "TC-057", "category": "Functionality", "page": "Scan Page",       "severity": "Medium", "description": "Verify URL input has correct type='url'", "input": "Inspect #url-input type", "expected": "type attribute equals 'url'"},
    "test_tc058_profile_scan_has_username_input":     {"id": "TC-058", "category": "Functionality", "page": "Scan Page",       "severity": "Critical","description": "Verify Profile scan page has username input", "input": "Navigate to /scan/profile", "expected": "#profile-username exists"},
    "test_tc059_profile_page_five_input_fields":      {"id": "TC-059", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify Profile page shows all 5 input fields", "input": "Navigate to /scan/profile", "expected": "All 5 profile input IDs present"},
    "test_tc060_profile_page_has_bio_textarea":       {"id": "TC-060", "category": "Functionality", "page": "Scan Page",       "severity": "Medium", "description": "Verify Profile page has bio textarea", "input": "Navigate to /scan/profile", "expected": "Textarea element found"},
    "test_tc061_scan_button_disabled_no_input":       {"id": "TC-061", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify scan button is disabled when no input provided", "input": "Navigate to /scan/text without typing", "expected": "Scan button has disabled attribute"},
    "test_tc062_scan_button_enabled_after_text":      {"id": "TC-062", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify scan button enabled after providing text input", "input": "Type text in textarea", "expected": "Scan button enabled (no disabled attr)"},
    "test_tc063_scan_button_enabled_after_url":       {"id": "TC-063", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify scan button enabled after entering URL", "input": "Type URL in input", "expected": "Scan button enabled"},
    "test_tc064_profile_scan_enabled_with_username":  {"id": "TC-064", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify profile scan enabled after entering username", "input": "Type @testbot in username", "expected": "Scan button enabled"},
    "test_tc065_info_card_shown":                      {"id": "TC-065", "category": "Functionality", "page": "Scan Page",       "severity": "Low",    "description": "Verify info card with AI model description is shown", "input": "Navigate to /scan/image", "expected": "glass-card with model info present"},
    "test_tc066_progress_animation_during_scan":      {"id": "TC-066", "category": "Functionality", "page": "Scan Page",       "severity": "High",   "description": "Verify progress animation appears during text scan", "input": "Submit text scan", "expected": "'Analyzing' or '%' indicator shown"},
    "test_tc067_scan_completes_redirects_to_result":  {"id": "TC-067", "category": "Functionality", "page": "Scan Page",       "severity": "Critical","description": "Verify scan completes and redirects to result page", "input": "Submit text scan, wait 5s", "expected": "URL contains /result"},
    "test_tc068_powered_by_subtitle":                  {"id": "TC-068", "category": "Functionality", "page": "Scan Page",       "severity": "Low",    "description": "Verify 'Powered by SocialShield AI' subtitle present", "input": "Navigate to /scan/image", "expected": "'SocialShield' or 'Powered by' text"},

    # ── Result Page ──
    "test_tc069_result_page_loads_with_header":       {"id": "TC-069", "category": "Functionality", "page": "Result Page",     "severity": "High",   "description": "Verify result page loads with 'Scan Result' header", "input": "Navigate to /result with cached data", "expected": "'Scan Result' or 'Result' in text"},
    "test_tc070_verdict_banner_displays":              {"id": "TC-070", "category": "Functionality", "page": "Result Page",     "severity": "Critical","description": "Verify verdict banner displays FAKE/REAL/SUSPICIOUS", "input": "Load result page with FAKE verdict", "expected": "'FAKE' text displayed"},
    "test_tc071_confidence_ring_renders":              {"id": "TC-071", "category": "Functionality", "page": "Result Page",     "severity": "High",   "description": "Verify confidence ring SVG renders", "input": "Load result page", "expected": "confidence-ring or SVG element exists"},
    "test_tc072_risk_level_indicator_shown":           {"id": "TC-072", "category": "Functionality", "page": "Result Page",     "severity": "High",   "description": "Verify risk level indicator is shown", "input": "Load result page", "expected": "'HIGH', 'MEDIUM', or 'LOW' with 'RISK'"},
    "test_tc073_fake_probability_displayed":           {"id": "TC-073", "category": "Functionality", "page": "Result Page",     "severity": "High",   "description": "Verify fake probability percentage is displayed", "input": "Load FAKE result", "expected": "'94.2%' or 'Fake' text present"},
    "test_tc074_real_probability_bar_renders":         {"id": "TC-074", "category": "Functionality", "page": "Result Page",     "severity": "Medium", "description": "Verify probability progress bars render", "input": "Load result page", "expected": "≥2 progress-bar-fill elements"},
    "test_tc075_ai_explanation_section":               {"id": "TC-075", "category": "Functionality", "page": "Result Page",     "severity": "High",   "description": "Verify AI Explanation section with bullet points", "input": "Load result page", "expected": "'AI Explanation' and bullet text present"},
    "test_tc076_technical_details_shown":              {"id": "TC-076", "category": "Functionality", "page": "Result Page",     "severity": "Medium", "description": "Verify Technical Details metadata table is shown", "input": "Load result page", "expected": "'Technical Details' or model name present"},
    "test_tc077_scan_id_displayed":                    {"id": "TC-077", "category": "Functionality", "page": "Result Page",     "severity": "Low",    "description": "Verify Scan ID is displayed at the bottom", "input": "Load result page", "expected": "'Scan ID' text present"},
    "test_tc078_back_button_navigates":                {"id": "TC-078", "category": "Functionality", "page": "Result Page",     "severity": "Medium", "description": "Verify '← Back' button is present", "input": "Load result page", "expected": "Button with 'Back' text exists"},
    "test_tc079_scan_again_button_navigates_home":    {"id": "TC-079", "category": "Functionality", "page": "Result Page",     "severity": "High",   "description": "Verify 'Scan Again' button navigates to home", "input": "Click 'Scan Again' button", "expected": "URL contains /home"},
    "test_tc080_verdict_color_matches_type":           {"id": "TC-080", "category": "Functionality", "page": "Result Page",     "severity": "Medium", "description": "Verify verdict color coding matches FAKE verdict", "input": "Load FAKE result", "expected": "'FAKE' shown with red (#FF3B3B) color"},

    # ── History Page ──
    "test_tc081_page_header_scan_history":             {"id": "TC-081", "category": "Functionality", "page": "History Page",    "severity": "High",   "description": "Verify page header shows 'Scan History'", "input": "Navigate to /history", "expected": "'Scan History' text present"},
    "test_tc082_subtitle_activity_log":                {"id": "TC-082", "category": "Functionality", "page": "History Page",    "severity": "Low",    "description": "Verify subtitle about scan activity log", "input": "Navigate to /history", "expected": "'activity log' text present"},
    "test_tc083_seven_filter_chips_render":            {"id": "TC-083", "category": "Functionality", "page": "History Page",    "severity": "High",   "description": "Verify all 7 filter chips render", "input": "Navigate to /history", "expected": "≥7 chip elements present"},
    "test_tc084_all_filter_active_by_default":        {"id": "TC-084", "category": "Functionality", "page": "History Page",    "severity": "Medium", "description": "Verify 'ALL' filter is active by default", "input": "Navigate to /history", "expected": "#filter-all has 'active' class"},
    "test_tc085_clicking_filter_changes_active":      {"id": "TC-085", "category": "Functionality", "page": "History Page",    "severity": "High",   "description": "Verify clicking filter chip changes active state", "input": "Click IMAGE filter", "expected": "IMAGE filter becomes active, ALL deactivated"},
    "test_tc086_history_items_render_with_icons":     {"id": "TC-086", "category": "Functionality", "page": "History Page",    "severity": "High",   "description": "Verify history items render with type icons", "input": "Navigate to /history", "expected": "history-item and icon-box elements present"},
    "test_tc087_history_item_shows_type_and_timestamp":{"id": "TC-087","category": "Functionality", "page": "History Page",    "severity": "Medium", "description": "Verify history items show media type and timestamp", "input": "Inspect first history item", "expected": "'Scan' text present in item"},
    "test_tc088_verdict_badges_display":               {"id": "TC-088", "category": "Functionality", "page": "History Page",    "severity": "High",   "description": "Verify verdict badges display correctly", "input": "Navigate to /history", "expected": "verdict-badge elements present"},
    "test_tc089_confidence_percentage_shown":          {"id": "TC-089", "category": "Functionality", "page": "History Page",    "severity": "Medium", "description": "Verify confidence percentage per item", "input": "Inspect first history item", "expected": "'%' in item text"},
    "test_tc090_delete_button_present":                {"id": "TC-090", "category": "Functionality", "page": "History Page",    "severity": "Medium", "description": "Verify delete button on history items", "input": "Inspect first history item", "expected": "Delete button element present"},
    "test_tc091_clicking_item_navigates_to_result":   {"id": "TC-091", "category": "Functionality", "page": "History Page",    "severity": "High",   "description": "Verify clicking history item navigates to result", "input": "Click first history item", "expected": "URL contains /result"},
    "test_tc092_empty_state_when_filtered":            {"id": "TC-092", "category": "Functionality", "page": "History Page",    "severity": "Medium", "description": "Verify empty state or items displayed", "input": "Navigate to /history", "expected": "Either items or 'No scans found' shown"},

    # ── Fraud Map Page ──
    "test_tc093_page_header_global_fraud_map":        {"id": "TC-093", "category": "Functionality", "page": "Fraud Map Page",  "severity": "High",   "description": "Verify 'Global Fraud Map' header", "input": "Navigate to /map", "expected": "'Global Fraud Map' or 'Fraud Map' text"},
    "test_tc094_four_summary_stat_cards":              {"id": "TC-094", "category": "Functionality", "page": "Fraud Map Page",  "severity": "High",   "description": "Verify 4 summary stat cards render", "input": "Navigate to /map", "expected": "≥4 stat-card elements"},
    "test_tc095_total_incidents_stat_value":           {"id": "TC-095", "category": "Functionality", "page": "Fraud Map Page",  "severity": "Medium", "description": "Verify 'Total Incidents' stat displays value", "input": "Navigate to /map", "expected": "'Total Incidents' and '12,655' present"},
    "test_tc096_map_container_renders":                {"id": "TC-096", "category": "Functionality", "page": "Fraud Map Page",  "severity": "Critical","description": "Verify map container renders", "input": "Navigate to /map", "expected": "map-container element present"},
    "test_tc097_nine_region_markers_present":          {"id": "TC-097", "category": "Functionality", "page": "Fraud Map Page",  "severity": "High",   "description": "Verify at least 9 region markers on the map", "input": "Navigate to /map", "expected": "≥7 region name labels found"},
    "test_tc098_clicking_region_shows_tooltip":       {"id": "TC-098", "category": "Functionality", "page": "Fraud Map Page",  "severity": "Medium", "description": "Verify clicking a region shows a tooltip", "input": "Click first region bubble", "expected": "Tooltip with Incidents/Trend/Risk shown"},
    "test_tc099_tooltip_shows_details":                {"id": "TC-099", "category": "Functionality", "page": "Fraud Map Page",  "severity": "Medium", "description": "Verify tooltip shows incidents, trend, risk", "input": "Click a region bubble", "expected": "Structured tooltip with details"},
    "test_tc100_risk_legend_present":                  {"id": "TC-100", "category": "Functionality", "page": "Fraud Map Page",  "severity": "Low",    "description": "Verify risk legend (HIGH/MEDIUM/LOW) present", "input": "Navigate to /map", "expected": "'HIGH', 'LOW', and 'RISK' text present"},
    "test_tc101_incident_breakdown_section":           {"id": "TC-101", "category": "Functionality", "page": "Fraud Map Page",  "severity": "High",   "description": "Verify 'Incident Type Breakdown' section", "input": "Navigate to /map", "expected": "'Incident Type Breakdown' text present"},
    "test_tc102_four_incident_type_bars":              {"id": "TC-102", "category": "Functionality", "page": "Fraud Map Page",  "severity": "Medium", "description": "Verify 4 incident type progress bars", "input": "Navigate to /map", "expected": "≥3 incident types and ≥4 progress bars"},

    # ── Settings Page ──
    "test_tc103_page_header_settings":                 {"id": "TC-103", "category": "Functionality", "page": "Settings Page",   "severity": "High",   "description": "Verify page header shows 'Settings'", "input": "Navigate to /settings", "expected": "'Settings' text present"},
    "test_tc104_profile_card_displays_user_info":     {"id": "TC-104", "category": "Functionality", "page": "Settings Page",   "severity": "High",   "description": "Verify profile card shows user info", "input": "Navigate to /settings", "expected": "Display name or email visible"},
    "test_tc105_edit_button_on_profile":               {"id": "TC-105", "category": "Functionality", "page": "Settings Page",   "severity": "Low",    "description": "Verify 'Edit' button on profile card", "input": "Navigate to /settings", "expected": "Button with 'Edit' text exists"},
    "test_tc106_pro_plan_badge_visible":               {"id": "TC-106", "category": "Functionality", "page": "Settings Page",   "severity": "Low",    "description": "Verify 'Pro Plan' badge is visible", "input": "Navigate to /settings", "expected": "'Pro Plan' text present"},
    "test_tc107_preferences_four_toggles":             {"id": "TC-107", "category": "Functionality", "page": "Settings Page",   "severity": "High",   "description": "Verify Preferences has ≥4 toggle switches", "input": "Navigate to /settings", "expected": "≥4 toggle-switch elements"},
    "test_tc108_dark_mode_on_by_default":              {"id": "TC-108", "category": "Functionality", "page": "Settings Page",   "severity": "Medium", "description": "Verify Dark Mode toggle is ON by default", "input": "Navigate to /settings", "expected": "#dark-mode checkbox is selected"},
    "test_tc109_toggle_switches_clickable":            {"id": "TC-109", "category": "Functionality", "page": "Settings Page",   "severity": "High",   "description": "Verify toggle switches change state on click", "input": "Click notifications toggle", "expected": "Toggle state flips"},
    "test_tc110_advanced_section_backend_url":         {"id": "TC-110", "category": "Functionality", "page": "Settings Page",   "severity": "Medium", "description": "Verify Advanced section shows Backend URL", "input": "Navigate to /settings", "expected": "'Backend URL' or 'Connected' text"},
    "test_tc111_ai_models_section_four_models":       {"id": "TC-111", "category": "Functionality", "page": "Settings Page",   "severity": "High",   "description": "Verify AI Models section lists 4 models", "input": "Navigate to /settings", "expected": "4 model names present"},
    "test_tc112_about_section_version":                {"id": "TC-112", "category": "Functionality", "page": "Settings Page",   "severity": "Low",    "description": "Verify About section shows Version 1.0.0", "input": "Navigate to /settings", "expected": "'Version' and '1.0.0' text"},
    "test_tc113_sign_out_button_danger_zone":          {"id": "TC-113", "category": "Functionality", "page": "Settings Page",   "severity": "High",   "description": "Verify 'Sign Out' in Danger Zone", "input": "Navigate to /settings", "expected": "#logout-btn with 'Sign Out' text"},
    "test_tc114_clear_history_button_danger_zone":    {"id": "TC-114", "category": "Functionality", "page": "Settings Page",   "severity": "Medium", "description": "Verify 'Clear History' in Danger Zone", "input": "Navigate to /settings", "expected": "'Danger Zone' and 'Clear History' text"},

    # ── Navigation & Layout ──
    "test_tc115_sidebar_renders_with_logo":            {"id": "TC-115", "category": "Functionality", "page": "Navigation",     "severity": "High",   "description": "Verify sidebar renders with SocialShield logo", "input": "Navigate to /home (desktop)", "expected": "sidebar and sidebar-logo elements present"},
    "test_tc116_four_nav_items_in_sidebar":            {"id": "TC-116", "category": "Functionality", "page": "Navigation",     "severity": "High",   "description": "Verify 4 navigation items in sidebar", "input": "Navigate to /home", "expected": "≥4 nav-item elements with labels"},
    "test_tc117_active_nav_item_highlighted":          {"id": "TC-117", "category": "Functionality", "page": "Navigation",     "severity": "Medium", "description": "Verify active nav item is highlighted", "input": "Navigate to /home", "expected": "nav-item.active element exists"},
    "test_tc118_mobile_nav_on_small_viewport":        {"id": "TC-118", "category": "Responsive",    "page": "Navigation",     "severity": "High",   "description": "Verify mobile bottom nav on 375×812 viewport", "input": "Set viewport 375×812", "expected": "mobile-nav visible"},
    "test_tc119_backend_status_dot_visible":           {"id": "TC-119", "category": "Functionality", "page": "Navigation",     "severity": "Medium", "description": "Verify backend status dot indicator", "input": "Navigate to /home", "expected": "Backend/Online/Offline text present"},
    "test_tc120_user_info_in_sidebar_footer":          {"id": "TC-120", "category": "Functionality", "page": "Navigation",     "severity": "Medium", "description": "Verify user info in sidebar footer", "input": "Navigate to /home", "expected": "'Signed in as' or user info text"},
    "test_tc121_sidebar_sign_out_functional":          {"id": "TC-121", "category": "Functionality", "page": "Navigation",     "severity": "High",   "description": "Verify sidebar Sign Out button present", "input": "Navigate to /home", "expected": "#sidebar-signout with 'Sign Out'"},
    "test_tc122_unknown_route_redirects":              {"id": "TC-122", "category": "Functionality", "page": "Navigation",     "severity": "High",   "description": "Verify unknown routes redirect", "input": "Navigate to /unknown-route", "expected": "URL doesn't contain 'unknown'"},

    # ── Vulnerability ──
    "test_tc123_xss_email_field_script_tags":          {"id": "TC-123", "category": "Vulnerability", "page": "Auth Page",       "severity": "Critical","description": "XSS via script tags in email field is sanitized", "input": "Enter <script>alert('XSS')</script> in email", "expected": "No alert dialog triggered"},
    "test_tc124_xss_text_scan_input":                  {"id": "TC-124", "category": "Vulnerability", "page": "Scan Page",       "severity": "Critical","description": "XSS in text scan textarea — no script execution", "input": "Enter <img onerror=alert('XSS')>", "expected": "No alert dialog triggered"},
    "test_tc125_xss_url_input_malicious":              {"id": "TC-125", "category": "Vulnerability", "page": "Scan Page",       "severity": "Critical","description": "XSS via malicious javascript: URL handled", "input": "Enter javascript:alert('XSS')", "expected": "No alert triggered, value stored as text"},
    "test_tc126_xss_profile_username_injection":      {"id": "TC-126", "category": "Vulnerability", "page": "Scan Page",       "severity": "Critical","description": "XSS via SVG onload in profile username prevented", "input": "Enter SVG onload payload", "expected": "No alert triggered"},
    "test_tc127_sql_injection_email_no_error":         {"id": "TC-127", "category": "Vulnerability", "page": "Auth Page",       "severity": "Critical","description": "SQL injection in email doesn't expose errors", "input": "Enter SQL payload in email", "expected": "No SQL/database errors in body text"},
    "test_tc128_html_injection_text_scan":              {"id": "TC-128", "category": "Vulnerability", "page": "Scan Page",       "severity": "High",   "description": "HTML injection doesn't manipulate DOM", "input": "Enter HTML tags in textarea", "expected": "No injected H1 in DOM"},
    "test_tc129_long_input_no_crash":                  {"id": "TC-129", "category": "Vulnerability", "page": "Scan Page",       "severity": "High",   "description": "10000+ character input doesn't crash page", "input": "Inject 10001-char string via JS", "expected": "Page remains functional"},
    "test_tc130_special_chars_in_password":             {"id": "TC-130", "category": "Vulnerability", "page": "Auth Page",       "severity": "Medium", "description": "Special characters in password field handled", "input": "Enter P@$$w0rd!#%^&*()", "expected": "Value accepted without error"},
    "test_tc131_csrf_tokens_not_in_dom":               {"id": "TC-131", "category": "Vulnerability", "page": "Home Page",       "severity": "High",   "description": "Auth tokens not exposed in visible DOM", "input": "Check body text on /home", "expected": "Token string not in page text"},
    "test_tc132_auth_token_cleared_on_logout":         {"id": "TC-132", "category": "Vulnerability", "page": "Settings Page",   "severity": "Critical","description": "Auth token cleared from localStorage on logout", "input": "Click Sign Out", "expected": "auth_token is null in localStorage"},

    # ── Responsive ──
    "test_tc133_mobile_viewport_renders":              {"id": "TC-133", "category": "Responsive",    "page": "Home Page",       "severity": "High",   "description": "App renders on mobile viewport (375×812)", "input": "Set viewport 375×812, go to /home", "expected": "Content visible, root ≤375px wide"},
    "test_tc134_tablet_viewport_layout":               {"id": "TC-134", "category": "Responsive",    "page": "Home Page",       "severity": "Medium", "description": "Layout adjusts on tablet viewport (768×1024)", "input": "Set viewport 768×1024", "expected": "Content renders correctly"},
    "test_tc135_desktop_viewport_sidebar_visible":    {"id": "TC-135", "category": "Responsive",    "page": "Home Page",       "severity": "High",   "description": "Sidebar visible on desktop (1440×900)", "input": "Set viewport 1440×900", "expected": "Sidebar element visible"},
    "test_tc136_mobile_nav_hidden_on_desktop":        {"id": "TC-136", "category": "Responsive",    "page": "Navigation",     "severity": "Medium", "description": "Mobile nav hidden on desktop viewport", "input": "Check at 1440×900", "expected": "mobile-nav display=none or height=0"},
    "test_tc137_sidebar_hidden_on_mobile":             {"id": "TC-137", "category": "Responsive",    "page": "Navigation",     "severity": "Medium", "description": "Sidebar hidden on mobile viewport", "input": "Set viewport 375×812", "expected": "sidebar display=none or width=0"},
    "test_tc138_scan_cards_stack_on_mobile":            {"id": "TC-138", "category": "Responsive",    "page": "Home Page",       "severity": "Medium", "description": "Scan cards stack vertically on mobile", "input": "Set viewport 375×812, check card positions", "expected": "Card2 Y > Card1 Y"},
    "test_tc139_auth_card_centered_all_viewports":    {"id": "TC-139", "category": "Responsive",    "page": "Auth Page",       "severity": "Medium", "description": "Auth card centered on all viewports", "input": "Check 375, 768, 1440 viewports", "expected": "Card center within 80px of viewport center"},
    "test_tc140_page_title_correct_all_pages":         {"id": "TC-140", "category": "Responsive",    "page": "All Pages",       "severity": "Medium", "description": "Page title 'SocialShield' on all navigated pages", "input": "Navigate to home, history, map, settings", "expected": "Title contains 'SocialShield' on each"},
}


# ─── Styling Constants ────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill(start_color="1B1B3A", end_color="1B1B3A", fill_type="solid")
HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="00D4FF")
PASS_FILL     = PatternFill(start_color="0A3D0A", end_color="0A3D0A", fill_type="solid")
PASS_FONT     = Font(name="Calibri", size=10, bold=True, color="06FFA5")
FAIL_FILL     = PatternFill(start_color="3D0A0A", end_color="3D0A0A", fill_type="solid")
FAIL_FONT     = Font(name="Calibri", size=10, bold=True, color="FF3B3B")
SKIP_FILL     = PatternFill(start_color="3D3D0A", end_color="3D3D0A", fill_type="solid")
SKIP_FONT     = Font(name="Calibri", size=10, bold=True, color="FFB800")
DATA_FONT     = Font(name="Calibri", size=10, color="FFFFFF")
DATA_FILL     = PatternFill(start_color="0D0D2B", end_color="0D0D2B", fill_type="solid")
ALT_FILL      = PatternFill(start_color="12122E", end_color="12122E", fill_type="solid")
THIN_BORDER   = Border(
    left=Side(style="thin", color="2A2A5A"),
    right=Side(style="thin", color="2A2A5A"),
    top=Side(style="thin", color="2A2A5A"),
    bottom=Side(style="thin", color="2A2A5A"),
)
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ─── Column Widths ────────────────────────────────────────────────────────────
COLUMNS = [
    ("S.No",              6),
    ("Test Case ID",     12),
    ("Category",         14),
    ("Page / Module",    16),
    ("Test Case Description", 50),
    ("Input / Action",   35),
    ("Expected Result",  35),
    ("Actual Result",    35),
    ("Status",           10),
    ("Execution Time",   14),
    ("Severity",         10),
    ("Error Details",    50),
]


def parse_junit_xml(xml_path):
    """Parse JUnit XML and return list of test result dicts."""
    tree = ET.parse(xml_path)
    root = tree.getroot()
    results = []

    for suite in root.iter("testsuite"):
        for case in suite.iter("testcase"):
            name = case.get("name", "")
            classname = case.get("classname", "")
            time_val = float(case.get("time", "0"))

            status = "PASS"
            error_msg = ""

            failure = case.find("failure")
            error = case.find("error")
            skipped = case.find("skipped")

            if failure is not None:
                status = "FAIL"
                error_msg = failure.get("message", "") or failure.text or ""
            elif error is not None:
                status = "FAIL"
                error_msg = error.get("message", "") or error.text or ""
            elif skipped is not None:
                status = "SKIP"
                error_msg = skipped.get("message", "") or ""

            results.append({
                "name": name,
                "classname": classname,
                "time": time_val,
                "status": status,
                "error": error_msg[:500],  # Truncate long errors
            })

    return results


def build_report(results, output_path):
    """Build a styled .xlsx report from parsed test results."""
    wb = Workbook()

    # ─── Sheet 1: Test Results ────────────────────────────────────────────
    ws = wb.active
    ws.title = "E2E Test Results"
    ws.sheet_properties.tabColor = "00D4FF"

    # Set column widths
    for col_idx, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row
    for col_idx, (header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, result in enumerate(results, 2):
        meta = TEST_METADATA.get(result["name"], {})
        sno = row_idx - 1
        tc_id = meta.get("id", f"TC-{sno:03d}")
        category = meta.get("category", "Functionality")
        page = meta.get("page", "Unknown")
        description = meta.get("description", result["name"].replace("_", " ").replace("test ", ""))
        input_action = meta.get("input", "—")
        expected = meta.get("expected", "—")
        status = result["status"]
        exec_time = f"{result['time']:.2f}s"
        severity = meta.get("severity", "Medium")
        error_details = result["error"] if status == "FAIL" else ("Skipped" if status == "SKIP" else "—")
        actual = f"{'PASSED' if status == 'PASS' else 'FAILED' if status == 'FAIL' else 'SKIPPED'} — {description}"

        row_fill = DATA_FILL if row_idx % 2 == 0 else ALT_FILL
        row_data = [sno, tc_id, category, page, description, input_action, expected, actual, status, exec_time, severity, error_details]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col_idx in (1, 2, 9, 10, 11) else LEFT_ALIGN

        # Color-code status cell
        status_cell = ws.cell(row=row_idx, column=9)
        if status == "PASS":
            status_cell.fill = PASS_FILL
            status_cell.font = PASS_FONT
        elif status == "FAIL":
            status_cell.fill = FAIL_FILL
            status_cell.font = FAIL_FONT
        elif status == "SKIP":
            status_cell.fill = SKIP_FILL
            status_cell.font = SKIP_FONT

        # Color-code severity
        severity_cell = ws.cell(row=row_idx, column=11)
        sev_colors = {"Critical": "FF3B3B", "High": "FFB800", "Medium": "00D4FF", "Low": "06FFA5"}
        severity_cell.font = Font(name="Calibri", size=10, bold=True, color=sev_colors.get(severity, "FFFFFF"))

    # ─── Sheet 2: Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_properties.tabColor = "06FFA5"

    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    pass_rate = (passed / total * 100) if total > 0 else 0
    total_time = sum(r["time"] for r in results)

    summary_data = [
        ("SocialShield E2E Test Report", ""),
        ("", ""),
        ("Report Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
        ("Application", "SocialShield — AI Deepfake & Fraud Detection"),
        ("Test Framework", "Selenium + Pytest (Python)"),
        ("Browser", "Google Chrome (Headless)"),
        ("", ""),
        ("RESULTS SUMMARY", ""),
        ("Total Test Cases", total),
        ("Passed ✅", passed),
        ("Failed ❌", failed),
        ("Skipped ⏭️", skipped),
        ("Pass Rate", f"{pass_rate:.1f}%"),
        ("Total Execution Time", f"{total_time:.1f}s"),
        ("", ""),
        ("CATEGORY BREAKDOWN", ""),
    ]

    # Category breakdown
    categories = {}
    for r in results:
        meta = TEST_METADATA.get(r["name"], {})
        cat = meta.get("category", "Functionality")
        if cat not in categories:
            categories[cat] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
        categories[cat]["total"] += 1
        categories[cat][r["status"].lower()] += 1

    for cat, counts in categories.items():
        summary_data.append((cat, f"{counts['pass']}/{counts['total']} passed ({counts['fail']} failed, {counts['skip']} skipped)"))

    summary_data.append(("", ""))
    summary_data.append(("PAGE BREAKDOWN", ""))

    # Page breakdown
    pages = {}
    for r in results:
        meta = TEST_METADATA.get(r["name"], {})
        page = meta.get("page", "Unknown")
        if page not in pages:
            pages[page] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
        pages[page]["total"] += 1
        pages[page][r["status"].lower()] += 1

    for page, counts in pages.items():
        summary_data.append((page, f"{counts['pass']}/{counts['total']} passed ({counts['fail']} failed, {counts['skip']} skipped)"))

    # Write summary
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)

        cell_a.fill = DATA_FILL
        cell_b.fill = DATA_FILL
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER

        if label in ("SocialShield E2E Test Report",):
            cell_a.font = Font(name="Calibri", size=16, bold=True, color="00D4FF")
            cell_a.fill = HEADER_FILL
            cell_b.fill = HEADER_FILL
        elif label in ("RESULTS SUMMARY", "CATEGORY BREAKDOWN", "PAGE BREAKDOWN"):
            cell_a.font = Font(name="Calibri", size=12, bold=True, color="FFB800")
            cell_a.fill = PatternFill(start_color="1A1A3E", end_color="1A1A3E", fill_type="solid")
            cell_b.fill = PatternFill(start_color="1A1A3E", end_color="1A1A3E", fill_type="solid")
        elif label == "Passed ✅":
            cell_b.font = PASS_FONT
        elif label == "Failed ❌":
            cell_b.font = FAIL_FONT
        elif label == "Skipped ⏭️":
            cell_b.font = SKIP_FONT
        elif label == "Pass Rate":
            color = "06FFA5" if pass_rate >= 80 else "FFB800" if pass_rate >= 50 else "FF3B3B"
            cell_b.font = Font(name="Calibri", size=14, bold=True, color=color)
        else:
            cell_a.font = Font(name="Calibri", size=10, bold=True, color="AAAACC")
            cell_b.font = DATA_FONT

    # Save
    wb.save(output_path)
    print(f"✅ Report saved: {output_path}")
    return output_path


def main():
    """Main entry point — parse XML and generate report."""
    os.makedirs(RESULTS_DIR, exist_ok=True)

    if not os.path.exists(XML_PATH):
        print(f"❌ JUnit XML not found at: {XML_PATH}")
        print("   Run tests first: pytest tests/ --junitxml=test-results/results.xml")
        sys.exit(1)

    print(f"📄 Parsing: {XML_PATH}")
    results = parse_junit_xml(XML_PATH)
    print(f"📊 Found {len(results)} test results")

    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%S")
    output_file = os.path.join(RESULTS_DIR, f"E2E_Test_Report_SocialShield_{timestamp}.xlsx")

    build_report(results, output_file)

    # Print summary
    total = len(results)
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] == "FAIL")
    skipped = sum(1 for r in results if r["status"] == "SKIP")
    pass_rate = (passed / total * 100) if total > 0 else 0

    print(f"\n{'='*50}")
    print(f"  SOCIALSHIELD E2E TEST SUMMARY")
    print(f"{'='*50}")
    print(f"  Total:   {total}")
    print(f"  Passed:  {passed} ✅")
    print(f"  Failed:  {failed} ❌")
    print(f"  Skipped: {skipped} ⏭️")
    print(f"  Rate:    {pass_rate:.1f}%")
    print(f"{'='*50}\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
